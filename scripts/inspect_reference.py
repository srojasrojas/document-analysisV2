from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = PROJECT_ROOT / "tmp" / "referencia.xlsx"
REPORTS_DIR = PROJECT_ROOT / "reports"
SUMMARY_PATH = REPORTS_DIR / "referencia_resumen.md"
SAMPLES_PATH = REPORTS_DIR / "referencia_operador_samples.jsonl"
OPERATOR_RE = re.compile(r"\boperador(?:a|es|as)?\b", re.IGNORECASE)


def _stringify(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def main() -> int:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Reference Excel not found: {EXCEL_PATH}")

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    samples: list[dict] = []
    lines: list[str] = ["# Resumen de referencia.xlsx", ""]

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        first_row = next(rows, None)
        headers = [_stringify(cell) or f"Columna {index + 1}" for index, cell in enumerate(first_row or [])]
        match_count = 0
        sheet_examples: list[dict] = []
        for row_number, row in enumerate(rows, start=2):
            values = [_stringify(cell) for cell in row]
            joined = " | ".join(value for value in values if value)
            if not OPERATOR_RE.search(joined):
                continue
            match_count += 1
            record = {
                "sheet": sheet_name,
                "row": row_number,
                "text": joined,
                "columns": {headers[index] if index < len(headers) else f"Columna {index + 1}": value for index, value in enumerate(values) if value},
            }
            samples.append(record)
            if len(sheet_examples) < 8:
                sheet_examples.append(record)

        lines.append(f"## Hoja: {sheet_name}")
        lines.append("")
        lines.append(f"- Columnas: {', '.join(headers) if headers else '(sin encabezados)'}")
        lines.append(f"- Filas con operador: {match_count}")
        if sheet_examples:
            lines.append("- Ejemplos:")
            for example in sheet_examples:
                text = example["text"][:500]
                lines.append(f"  - Fila {example['row']}: {text}")
        lines.append("")

    SUMMARY_PATH.write_text("\n".join(lines), encoding="utf-8")
    with SAMPLES_PATH.open("w", encoding="utf-8") as handle:
        for sample in samples:
            handle.write(json.dumps(sample, ensure_ascii=False))
            handle.write("\n")

    print(f"Wrote {SUMMARY_PATH.relative_to(PROJECT_ROOT)}")
    print(f"Wrote {SAMPLES_PATH.relative_to(PROJECT_ROOT)} ({len(samples)} samples)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())